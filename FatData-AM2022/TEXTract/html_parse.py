# -*- coding: utf-8 -*-

from bs4 import BeautifulSoup
import os
import json
from openpyxl import load_workbook,Workbook

def ifskip(string):
    flag=0
    stop=['abbre','Bib','Ack','author','affiliation','corresponding', \
          'editor','right','info','cite','url']
    for sw in stop:
        if sw in string:
            flag=1
            break
    
    return flag

def ifskip_section(string):
    flag=0
    stop=['abbreviation','references','acknow','author information','editor information', \
          'rights and permissions','copyright information','about this paper','abstract',\
          'additional information','ethics','funding','notes','supplementary','about this article',\
          'avail']
    string=string.lower()
    for sw in stop:
        if sw in string:
            flag=1
            break
    
    return flag
   
def parse_paragraph(doc):
    text=doc.text.replace('\n','')
    return text

def parse_ol(doc):
    idlst=[]
    textlst=[]
    for i, child in enumerate(doc.children):
        if not (child.name is None):
            idlst.append(child.name)
            textlst.append(child.text.replace('\n',''))
    return idlst,textlst
    
def parse_section(doc):
    textlst=[]
    idlst=[] 
    skip=[]
    title=''
    for i, child in enumerate(doc.children):
        # print(i,child.name)
        # if (hasattr(child,'attrs')) and  ('id' in child.attrs.keys()):
        #     if ifskip(child.attrs['id']):
        #         continue
        #     print(child.attrs['id'])
        if child.name=='div':
            ids=[]
            texts=[]
            if 'class' in child.attrs.keys():
                if child.attrs['class'][0]=='c-article-equation__number':
                    ids.append('eq_num')
                    texts.append(child.text)
                elif child.attrs['class'][0]=='c-article-equation':
                    ids.append('eq')
                    texts.append(child.text)   
                else:
                    ids,texts=parse_section(child)
            else:
                ids,texts=parse_section(child)
            idlst.extend(ids)
            textlst.extend(texts)
        elif child.name=='h1' or child.name=='h2' or child.name=='h3' or \
             child.name=='h4' or child.name=='h5' or child.name=='h6':
                 title=child.text
                 idlst.append(child.name)
                 textlst.append(title)                 
        elif child.name=='p':
            texts=parse_paragraph(child)
            idlst.append('p')
            textlst.append(texts)
        elif child.name=='ol':
            ids,texts=parse_ol(child)
            idlst.extend(ids)
            textlst.extend(texts)            
    return idlst,textlst

def parse_abstract(abst):
    abstract=''
    keywords=[]
    tmp=abst.find_all('div',class_="c-article-section__content")
    para=tmp[0].find('p')
    abstract=para.text.replace('\n','')
    kws=tmp[0].find_all('li',class_='c-article-subject-list__subject')
    for k in kws:
        keywords.append(k.text)
    
    return abstract,keywords

def parse_doc(doc):
    
    idlst=[]
    textlst=[]
    title=''
    abstract=''
    keywords=[]
    # get title
    tmp=doc.find_all('h1',class_='c-article-title')
    if len(tmp)==1:
        title=tmp[0].text
    else:
        title=tmp[0].text
        print("Warning: multi title")
    
     
    # get section
    tmp=doc.find_all('section')
    for sec in tmp:
        #print(sec.attrs['data-title'])
        if ('data-title' in sec.attrs.keys())and(not ifskip_section(sec.attrs['data-title'])):
            tmp_id,tmp_text=parse_section(sec)
            idlst.append(tmp_id)
            textlst.append(tmp_text)
        elif ('data-title' in sec.attrs.keys())and('Abstract' in sec.attrs['data-title']):
            abstract,keywords=parse_abstract(sec)
    
        
    return title,abstract,keywords,idlst,textlst


def section_struct(ids,texts):
    sec={'sec_title':'','content':[]}
    minid=100
    record=[]
    ct=0
    for i in range(0,len(ids)):
        if ids[i][0]=='h':
            ct=ct+1
            current=int(ids[i][1:])
            if current<minid:
                minid=current


                
    for i in range(0,len(ids)):
        if ids[i][0]=='h':
            current=int(ids[i][1:])
            if current==minid:
                record.append(i)

    if len(record)==1:
        for i in range(0,len(ids)):
            if (ids[i][0]=='h')and(int(ids[i][1:])==minid):
                sec['sec_title']=texts[i]
            elif (ids[i][0]=='h')and(int(ids[i][1:])>minid):
                tmp=section_struct(ids[i:],texts[i:])
                sec['content'].append(tmp['content'])
                return sec
            else:
                sec['content'].append(texts[i])
        return sec
    else:            
        for i in range(0,len(record)-1):
            sec['content'].append(section_struct(ids[record[i]:record[i+1]],texts[record[i]:record[i+1]]))
        sec['content'].append(section_struct(ids[record[-1]:],texts[record[-1]:]))
    return sec


def doc_struct(title,abstract,keywords,ids,texts,path,file,doi):
    doc={}
    doc['title']=title
    doc['abstract']=abstract
    doc['keywords']=keywords
    doc['path']=path
    doc['file']=file
    doc['doi']=doi
    doc['content']=[]
    for i in range(0,len(ids)):
        doc['content'].append(section_struct(ids[i],texts[i]))
    return doc

def get_doi_dict():
    doi_dict={}
    wospath="D:/Tsinghua/Project/Fatigue Data Framework/search data/wos/20220917/"
    fileName="AM_fatigue_adddoi.xlsx"
    sheetName="savedrecs"
    wb=load_workbook(filename=wospath+fileName)
    ws=wb[sheetName]
    nrow=ws.max_row
    ncol=ws.max_column
    for i in range(1,ncol+1):
        if ws.cell(1,i).value=='DOI':
            col=i
            break
    for i in range(2,nrow+1):
        doi=ws.cell(i,col).value.strip('\n')
        doi_trans=doi.replace('/','_').replace(':','_')
        doi_dict[doi_trans]=doi

    return doi_dict

path=''
txtpath=''
logpath=txtpath+'parse.log'
doi_dict=get_doi_dict()
filelist=os.listdir(path)
doc_dict={}
iftxt=0

log=open(logpath,'w')
for file in filelist:
    leng=len(file)
    if leng>5 and file[leng-5:leng]=='.html':    
        filename=file[0:leng-5]
        doi=doi_dict[filename]
        print(file)
        log.write(file)
        try:
            f=open(path+file,'r',encoding='utf-8')
            doc=BeautifulSoup(f,"lxml")
            f.close()
            title,abstract,keywords,ids,texts=parse_doc(doc)
            ifsuccess=1
        except:
            ifsuccess=0
        if iftxt:
            fout=open(txtpath+filename+'.txt','w',encoding='utf-8')    
        if ifsuccess:
            doc_dict[doi]=doc_struct(title,abstract,keywords,ids,texts,path,file,doi)
            log.write('\n')
            if iftxt:
                fout.write('E:/Data/Literature Data/AM fatigue/'+filename+'.pdf\n')
                fout.write(title+'\n')
                fout.write('Abstract\n')
                fout.write(abstract+'\n')
                fout.write('Keywords\n')
                for kw in keywords:
                    fout.write(kw+', ')
                fout.write('\n')
                for i in range(0,len(texts)):
                    fout.write('[Section %d]\n'%(i+1))
                    for j in range(0,len(texts[i])):
                        fout.write('['+ids[i][j]+']'+'\n')
                        fout.write(texts[i][j]+'\n')
        else:
            if iftxt:
                fout.write(file+' PRASE ERROR\n')
            print("PRASE ERROR")
            log.write(' PRASE ERROR\n')
        if iftxt:
            fout.close()
log.close()

# f=open(path+'data.json','w',encoding='utf-8')
# json.dump(doc_dict,f)
# f.close()